#la libreria openpyxl nos permite trabajar con archivos excel desde Python de forma sencilla

import openpyxl


book = openpyxl.Workbook() #creamos un libro de excel nuevo

sheet = book.active

sheet["A1"] = "Product" # pyright: ignore[reportOptionalSubscript] #escribimos en la celda A1
sheet["B1"] = "Price"  # type: ignore #escribimos en la celda B1
sheet["C1"] = "quantity"  # type: ignore #escribimos en la celda C1

#Introducir datos en las filas siguientes
sheet["A2"] = "Laptop" # type: ignore
sheet["B2"] = 800 # type: ignore
sheet["C2"] = 5 # type: ignore

#Guardar el archivo excel
book.save("Python/especializado/manejo_excel/inventario.xlsx")

#creemos datos de ventas 

sheet["A1"] = "Product"  # type: ignore
sheet["B1"] = "Price"  # type: ignore
sheet["C1"] = "Quantity"  # type: ignore
sheet["A1"].font = openpyxl.styles.Font(bold=True)  # type: ignore
sheet["B1"].font = openpyxl.styles.Font(bold=True)  # type: ignore
sheet["C1"].font = openpyxl.styles.Font(bold=True, color="FF0000", size=32)  # type: ignore

sales = [
    ["Laptop", 800, 5],
    ["Mouse", 50, 20],
    ["Keyboard", 100, 10],
    ["Monitor", 200, 5],
    ["Printer", 150, 7],
]

#agregar todos los datos a la hoja de excel
for row in sales:
    sheet.append(row) #Lo guarda en la siguiente fila disponible

#guardar el archivo excel
book.save("Python/especializado/manejo_excel/sales.xlsx")


#Ver el contenido de un archivo excel existente

#cargamos el archivo excel
newBook = openpyxl.load_workbook("Python/especializado/manejo_excel/inventario.xlsx") #cargar un archivo excel existente

#obtenemos la hoja activa
newSheet = newBook.active
newSheet.title = "Inventory"  # type: ignore

#iteramos sobre las filas para obtener los valores
for row in newSheet.iter_rows():
  #iteramos sobre las celdas de cada fila para obtener los valores
    for cell in row:
      #imprimimos el valor de la celda
        print(cell.value, end=" ")
    print()
    
#Ejercicio 1: "Mi Primera Lista de Compras" (Básico)

""" Objetivo: Practicar la creación de un libro, selección de hoja, escritura en celdas específicas y guardado.

Instrucciones:

Crea un script en Python que importe Workbook de openpyxl.

Crea un nuevo libro y selecciona la hoja activa.

Cambia el título de la hoja a "Compras".

En la celda A1 escribe "Artículo" y en B1 escribe "Cantidad".

Agrega 3 artículos de tu elección en las filas siguientes (ej: "Manzanas" | 5).

Guarda el archivo con el nombre compras_semana.xlsx. """

#Solucion 1
secondBook = openpyxl.Workbook()
secondSheet = secondBook.active
secondSheet.title = "Compras"

secondSheet["A1"] = "Articulo"
secondBook["B1"] = "Cantidad"

articles = [
    ["Manzanas", 5],
    ["Peras", 10],
    ["Bananas", 15],
]

for row in articles:
    secondSheet.append(row)

secondBook.save("Python/especializado/manejo_excel/compras_semana.xlsx")

#Ejercicio 2: "El Generador de Tablas" (Lógica de Programación + Excel)

""" Objetivo: Integrar lógica de Python (bucles anidados) con la escritura en celdas.

Instrucciones:

Crea un libro nuevo.

El objetivo es crear una tabla de multiplicar del 1 al 10 en Excel.

Usa dos bucles for anidados (uno para filas i de 1 a 10, otro para columnas j de 1 a 10).

En cada celda, el valor debe ser el resultado de i * j.

Pista: Recuerda que puedes acceder a las celdas usando sheet.cell(row=i, column=j).value = resultado

Guarda el archivo como tabla_multiplicar.xlsx. """

book = openpyxl.Workbook()
sheet = book.active

for i in range(1,11):
  for j in range(1,11):
    result = i * j
    sheet.cell(row=i, column=j).value = result
    
book.save('Python/especializado/manejo_excel/tablas.xlsx')